import csv
import datetime
import json
import os

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

fields = ['Player', 'Total Points', 'Total Points Odds', 'Total Points + Assists + Rebounds',
          'Total Points + Assists + Rebounds Odds', 'Total Blocks + Steals', 'Total Blocks + Steals Odds',
          'Total 3pt Field Goals', 'Total 3pt Field Goals Odds']
headers = ["Player", "WH Points", "WH Point Odds", "WH PAR", "WH PAR Odds", "WH BS", "WH BS Odds", "WH 3PT",
           "WH 3PT Odds"]
test = False
debug = False


def main():
    logo()
    filename = f"WH NBA Betting Data {datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S')}.csv"
    print("Fetching highlights...")
    print("Output file", filename)
    with open(filename, "w", newline='') as f:
        csv.DictWriter(f, fieldnames=headers).writeheader()
    if test:
        with open('out.json') as outfile:
            res = outfile.read()
    else:
        res = requests.get('https://www.williamhill.com/us/va/bet/api/v2/events/highlights').text
    highlights = json.loads(res)
    if debug:
        print("highlights", json.dumps(highlights, indent=4))
    for sport in highlights:
        if sport['sportId'] == 'basketball':
            for comp in sport['competitions']:
                for event in comp['events']:
                    print('EventID', event["id"], event["name"])
                    if debug:
                        print("Event", json.dumps(event, indent=4))
                    if test:
                        with open('event.json') as efile:
                            res = efile.read()
                    else:
                        res = requests.get(f'https://www.williamhill.com/us/va/bet/api/v2/events/{event["id"]}').text
                    ev = json.loads(res)
                    data = {}
                    for market in ev['markets']:
                        if "| |" in market['name'] and "Half" not in market['name'] and "Quarter" not in market[
                            'name'] and "|Total Points| " not in market['name']:
                            try:
                                player = market['name'].strip().split("| |")[0][1:]
                                gametype = market['name'].strip().split("| |")[1][:-1]
                                # print(json.dumps(market, indent=4))
                                if player not in data.keys():
                                    data[player] = {}
                                data[player][f"{gametype}"] = market['line']
                                for sel in market['selections']:
                                    if sel['name'] == '|Over|':
                                        data[player][f"{gametype} Odds"] = sel['price']['a']
                                        break
                                # break
                            except:
                                pass
                    print(json.dumps(data, indent=4))
                    games = []
                    for x in data.keys():
                        if len(str(data[x])) > 5:
                            game = json.loads(json.dumps(data[x]))
                            game["Player"] = str(x)
                            # print(json.dumps(game, indent=4))
                            games.append(game)
                        else:
                            print(f"Empty player found {x} {data[x]}")
                    with open(filename, "a", newline='') as f:
                        w = csv.DictWriter(f, fieldnames=fields, extrasaction='ignore')
                        w.writerows(games)
                    if test:
                        break
                if test:
                    break
            if test:
                break
    print("Converting CSV to XSLX")
    cvrt(filename)
    input(f"Done!! Press any key. Output written to file {filename}")


def cvrt(filename):
    wb = Workbook()
    worksheet = wb.active
    with open(filename, 'r', encoding='utf8') as f:
        rows = [row for row in csv.reader(f)]
    column_widths = []
    for row in rows:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(cell) > column_widths[i]:
                    column_widths[i] = len(cell)
            else:
                column_widths += [len(cell)]
    for i, column_width in enumerate(column_widths):
        worksheet.column_dimensions[get_column_letter(i + 1)].width = column_width + 1

    for row in rows:
        worksheet.append(row)
    for col in worksheet.columns:
        for cell in col:
            cell.alignment = Alignment(horizontal='center')
    wb.save(filename.replace('.csv', '.xlsx'))


def logo():
    os.system('color 0a')
    print("""
     _       __ _  __ __ _                      __  __ _  __ __
    | |     / /(_)/ // /(_)____ _ ____ ___     / / / /(_)/ // /
    | | /| / // // // // // __ `// __ `__ \   / /_/ // // // / 
    | |/ |/ // // // // // /_/ // / / / / /  / __  // // // /  
    |__/|__//_//_//_//_/ \__,_//_/ /_/ /_/  /_/ /_//_//_//_/   
==================================================================
            WilliamHill.com betting odds scraper by:
                  github.com/evilgenius786
==================================================================
[+] Without browser
[+] Efficient and fast
[+] Works with API
[+] Output in XLSX and CSV
__________________________________________________________________                                                     
""")


if __name__ == '__main__':
    main()
